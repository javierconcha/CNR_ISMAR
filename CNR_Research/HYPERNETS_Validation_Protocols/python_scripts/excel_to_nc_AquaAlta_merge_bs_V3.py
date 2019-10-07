# -*- coding: utf-8 -*-
"""
Created on Mon Nov 20 16:10:50 2017
@author: Marco
Modified by Javier A. Concha on September 18 2019
"""
"""
Created on Mon Nov 20 16:10:50 2017
Function to create a netcdf file with the in situ data from AERONET-OC using V3.
Example how to call the function:
    python excel_to_nc_AquaAlta_merge_newsite.py -sd 2016 01 01 -ed 2016 12 31 -s Venise
    
Note: The following subfolder should be created in advance:
    - temp_file_15/
    - temp_file_20/
    - excel_file/
    - netcdf_file/
"""
"""
This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.
"""

import xlrd
import sys
import os
import numpy as np
from netCDF4 import Dataset
import csv
import openpyxl
import xlwt
import argparse
import wget

from collections import OrderedDict
parser= argparse.ArgumentParser()
parser.add_argument("-s", "--station" , nargs=1, help="The Aeronet OC station", required=True)
parser.add_argument("-sd", "--sdate", nargs=3, help="Three integers: start year, start month and start day ",required=True)
parser.add_argument("-ed", "--edate", nargs=3, help="Three integers: end year, end month and end day",required=True)
#parser.add_argument("-l", "--level", nargs=1, help="Three integers: end year, end month and end day",required=True)

args=parser.parse_args()
station=args.station[0]
starty=args.sdate[0]
startm=args.sdate[1]
startd=args.sdate[2]
endy=args.edate[0]
endm=args.edate[1]
endd=args.edate[2]
#level=args.level[0]
if int(startm)>12 or int(startm)<1:
    sys.exit("Non valid month")
if int(starty)%4==0:
    if int(startd)<1 or int(startd)>366:
        sys.exit("Non valid day for the start year")
else:
    if int(startd)<1 or int(startd)>365:
        sys.exit("Non valid day for the start year")

if int(endy)%4==0:
    if int(endd)<1 or int(endd)>366:
        sys.exit("Non valid day for the end year")
else:
    if int(endd)<1 or int(endd)>365:
        sys.exit("Non valid day for the end year")

station_list=['Abu_Al_Bukhoosh','Blyth_NOAH','COVE_SEAPRISM', 'Gageocho_Station','Galata_Platform','Gloria', 'GOT_Seaprism',
              'Gustav_Dalen_Tower','Helsinki_Lighthouse','Ieodo_Station','KAUST_Campus','Lake_Erie','LISCO','Lucinda',
              'MVCO', 'Palgrunden', 'Socheongcho','Thornton_C-power','USC_SEAPRISM','USC_SEAPRISM_2','Venise' ,'WaveCIS_Site_CSI_6','Zeebrugge-MOW1']
station_upper=[]
for s in station_list:
    station_upper.append(s.upper())
if station.upper() not in station_upper:
    sys.exit('The station does not exists, the stations are '+str(station_list)[1:-1])
ind=station_upper.index(station.upper())

book = xlwt.Workbook()
link_15='https://aeronet.gsfc.nasa.gov/cgi-bin//print_web_data_v3?site='+station_list[ind]+'&year='+starty+'&month='+startm+'&day='+startd+'&year2='+endy+'&month2='+endm+'&day2='+endd+'&LWN15=1&AVG=10'
link_20='https://aeronet.gsfc.nasa.gov/cgi-bin//print_web_data_v3?site='+station_list[ind]+'&year='+starty+'&month='+startm+'&day='+startd+'&year2='+endy+'&month2='+endm+'&day2='+endd+'&LWN20=1&AVG=10'
if len(startm)==1:
    startm='0'+startm
if len(endm)==1:
    endm='0'+endm
if len(startd)==1:
    startd='0'+startd


if len(endd)==1:
    endd='0'+endd

filename1_15=station_list[ind]+'_15V3_'+starty+startm+startd+'_'+endy+endm+endd+'.txt'
filename2_15=filename1_15[:-4]+'.xlsx'
filename1_20=station_list[ind]+'_20V3_'+starty+startm+startd+'_'+endy+endm+endd+'.txt'
filename2_20=filename1_20[:-4]+'.xlsx'
if os.path.isfile(filename2_15):
    sys.exit('Data already downloaded for the chosen period')

filename_15=wget.download(link_15, 'temp_file_15/')
filename_20=wget.download(link_20, 'temp_file_20/')
os.rename(filename_15, filename1_15)
os.rename(filename_20, filename1_20)
input_file_15 = filename1_15
file_excel_15 = 'excel_file/'+filename2_15
input_file_20 = filename1_20
file_excel_20 = 'excel_file/'+filename2_20
wb = openpyxl.Workbook()
ws = wb.worksheets[0]
with open(input_file_15, 'r') as data:
    reader = csv.reader(data,  delimiter=',')
    for row in reader:
        ws.append(row)
os.remove(filename1_15)
wb.save(file_excel_15)
wb = openpyxl.Workbook()
ws = wb.worksheets[0]
with open(input_file_20, 'r') as data:
    reader = csv.reader(data,  delimiter=',')
    for row in reader:
        ws.append(row)
os.remove(filename1_20)
wb.save(file_excel_20)

#file_excel='AquaAlta_15_201211_20171230.xlsx'
workbook = xlrd.open_workbook(file_excel_15)
worksheet0 = workbook.sheet_by_index(0)
data_15=[]
for row in range(worksheet0.nrows):
    data_15.append(worksheet0.row_values(row))
if len(data_15)==4:
    sys.exit('No data for the chosen period')
header=data_15[7]
head=header[-37:-9]
nw=len(head)


nwe=0
for head_el in head:
    if head_el.split('_')[-1]=='Empty':
        nwe+=1
data_15=data_15[8:-1]
data_15=np.array(data_15)
for j in range(len(data_15[:,-1])):
    data_15[j,-1]=data_15[j,-1][:-4]
date_15=data_15[:,1]
time_day_15=data_15[:,2]
j_day_15=data_15[:,3]
#julian_day=data[:,2]
time_d_15=[]
for (d, t) in zip(date_15, time_day_15):
    time_d_15.append(d+' '+t)
time_d_15=np.array(time_d_15)
workbook = xlrd.open_workbook(file_excel_20)
worksheet0 = workbook.sheet_by_index(0)
data_20=[]
for row in range(worksheet0.nrows):
    data_20.append(worksheet0.row_values(row))
if len(data_20)<8:
    print('Only level 1.5 data')
    data=data_15
    time_d=time_d_15
    j_day=j_day_15
    level_flag=np.zeros(len(time_d))
    level_flag=1.5

else:
    header=data_20[7]
    data_20=data_20[8:-1]
    data_20=np.array(data_20)
    for j in range(len(data_20[:,-1])):
        data_20[j,-1]=data_20[j,-1][:-4]
    date_20=data_20[:,1]
    time_day_20=data_20[:,2]
    j_day_20=data_20[:,3]
    #julian_day=data[:,2]
    time_d_20=[]
    for (d, t) in zip(date_20, time_day_20):
        time_d_20.append(d+' '+t)
    time_d_20=np.array(time_d_20)
    date_20_split=date_20[-1]
    index_20=len(date_20)

    index_split=np.where(date_15==date_20_split)[0][0]
    data=np.concatenate((data_20[:,:], data_15[index_split+1:,:]))
    time_d=np.concatenate((time_d_20, time_d_15[index_split+1:]))
    j_day=np.concatenate((j_day_20, j_day_15[index_split+1:]))
    level_flag=np.zeros(len(time_d))
    level_flag[:index_20]=2.0
    level_flag[index_20:]=1.5

filename='netcdf_file/'+filename1_20[:-3]+'nc'
nc_f=Dataset(filename, 'w', format='NETCDF4')
nc_f.createDimension('Time', len(time_d))
nc_f.createDimension('Central_wavelenghts', nw-nwe)
time=nc_f.createVariable('Time', 'str', ('Time', ))
time[:]=time_d

l_f=nc_f.createVariable('Level', 'f4', ('Time',))
l_f[:]=level_flag
j_d=nc_f.createVariable('Julian_day', str, ('Time',))
j_d[:]=j_day

#ins=nc_f.createVariable('Instrument_number', 'i4', ('Time', ))
[waves,sol_zen, sol_azi, Lt_mean, Lt_std_dev, Lt_min_rel, Li_mean, Li_std_dev, AOT, OOT, ROT, Lw, Lw_Q, Lwn, Lwn_fonQ]=[None]*15
variable_list=['Solar_zenith', 'Solar_azimuth', 'Lt_mean', 'Lt_standard_deviation',
               'Lt_min_rel', 'Li_mean', 'Li_standard_deviation', 'RHO', 'AOT', 'OOT', 'ROD', 'NO2OD', 'Water Vapour']

variable_list2=['Lw', 'Lw_Q', 'Lwn', 'Lwn_fonQ', 'Lwn_IOP' ,'Exact_wavelengths']
corr_dict=OrderedDict({})

for ind,key in enumerate(variable_list):
    startind=5+ind*nw
    endind=5+ind*nw+nw-nwe
    corr_dict[key]=data[:,startind:endind]

final=endind+6+nwe
for ind,key in enumerate(variable_list2):
    startind=final+ind*nw
    endind=final+ind*nw+(nw-nwe)  
    if key=="Exact_wavelengths":
        startind+=1
        endind+=1
        
    corr_dict[key]=data[:,startind:endind] 


    



for var in corr_dict:
    nc_v=nc_f.createVariable(var, 'f4', ('Time', 'Central_wavelenghts', ))
    if var=='Exact_wavelengths':
        vv=np.array(corr_dict[var]).astype('float')
        vv[np.where(vv!=-999.)]=vv[np.where(vv!=-999.)]*1000.
        nc_v[:]=vv

    else:
        nc_v[:]=corr_dict[var]
    
    
variable_list3=['Pressure', 'Wind_speed', 'CHL-A', 'Total_precipitable_water', 'Total_NO2', 'Total_Ozone']
for ind,v in enumerate(variable_list3):
    var_nc=nc_f.createVariable(v, 'f4', ('Time', ))
    var_nc[:]=data[:,final-(ind+1)]

nc_f.close()
